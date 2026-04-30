"""
Gera arquivos mock (.xlsx) para todos os módulos da Central de Relatórios.
Baseado na leitura exata do código de central_relatorios.py.

Executar a partir do diretório backend/:
    python tests/gerar_mocks.py
"""

import os
from datetime import date, timedelta
import openpyxl
from openpyxl import Workbook
import pandas as pd

FIXTURES = os.path.join(os.path.dirname(__file__), 'fixtures')
os.makedirs(FIXTURES, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# 1. PEDIDOS
# Detecção automática de aba com colunas: Data, Expedição, Depositante, Ordem de Saída
# ─────────────────────────────────────────────────────────────────────────────
def mock_pedidos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    ws.append(['Data', 'Expedição', 'Depositante', 'Ordem de Saída'])
    today = date.today()
    clientes = ['ADITUS', 'GSK', 'IPSEN', 'BIOGEN', 'YELUM']
    for i in range(10):
        inicio = today - timedelta(days=30 - i)
        fim = inicio + timedelta(days=i % 3)  # 0,1,2 dias
        cliente = clientes[i % len(clientes)]
        ws.append([
            inicio.strftime('%d/%m/%Y'),
            fim.strftime('%d/%m/%Y'),
            cliente,
            f'OS-{10000 + i}'
        ])
    path = os.path.join(FIXTURES, 'mock_pedidos.xlsx')
    wb.save(path)
    print(f"✅ mock_pedidos.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 2. FRETES
# Precisa de: aba com coluna 'valor frete' (+ Remetente, Peso Taxado, Volume,
# frete peso, itr, outros), aba RESCOM (remetente, valor), aba PORTADORES (remetente, valor parceiro)
# ─────────────────────────────────────────────────────────────────────────────
def mock_fretes():
    wb = Workbook()
    # Aba principal de EMBARQUES
    ws_emb = wb.active
    ws_emb.title = "EMBARQUES"
    ws_emb.append(['Remetente', 'valor frete', 'Peso Taxado', 'Volume',
                   'frete peso', 'itr', 'outros'])
    remetentes = ['ADITUS COMERCIAL LTDA', 'GSK BRASIL LTDA',
                  'IPSEN BIOFARMACEUTICA', 'BIOGEN LTDA', 'YELUM SEGURADORA']
    for i, rem in enumerate(remetentes):
        ws_emb.append([rem, 1500.0 + i * 200, 50.0 + i * 10, 2.0 + i * 0.5,
                       3.0 + i * 0.2, 20.0 + i * 5, 10.0 + i])

    # Aba RESCOM
    ws_res = wb.create_sheet('RESCOM')
    ws_res.append(['remetente', 'valor'])
    for i, rem in enumerate(remetentes):
        ws_res.append([rem, 300.0 + i * 50])

    # Aba PORTADORES
    ws_port = wb.create_sheet('PORTADORES')
    ws_port.append(['remetente', 'valor parceiro'])
    for i, rem in enumerate(remetentes):
        ws_port.append([rem, 150.0 + i * 25])

    path = os.path.join(FIXTURES, 'mock_fretes.xlsx')
    wb.save(path)
    print(f"✅ mock_fretes.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 3. ARMAZENAGEM
# Colunas: Emissão (DD/MM/YYYY), Cliente, Valor Principal
# ─────────────────────────────────────────────────────────────────────────────
def mock_armazenagem():
    wb = Workbook()
    ws = wb.active
    ws.title = "Armazenagem"
    ws.append(['Emissão', 'Cliente', 'Valor Principal'])
    clientes = ['ADITUS', 'GSK', 'IPSEN', 'BIOGEN', 'YELUM', 'CELLTRION']
    today = date.today()
    mes_atual = today.replace(day=1)
    for i, cli in enumerate(clientes):
        # Algumas datas no mês atual
        d = mes_atual + timedelta(days=i * 3)
        ws.append([d.strftime('%d/%m/%Y'), cli, 5000.0 + i * 1000])
    # Um cliente com pagamento fracionado (dois meses)
    mes_ant = (mes_atual - timedelta(days=15)).replace(day=1)
    ws.append([mes_ant.strftime('%d/%m/%Y'), 'ADITUS', 2500.0])

    path = os.path.join(FIXTURES, 'mock_armazenagem.xlsx')
    wb.save(path)
    print(f"✅ mock_armazenagem.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 4. ESTOQUE — CARGA INICIAL
# Abas por cliente. Header nas linhas 0-3 (4 linhas).
# Linha 4+: col0=Código, col1=(qualquer), col2=Descrição, col3..col5=(qualquer),
#           col6=Saldo, col7..col8=(qualquer), col9=Reservado,
#           col10=(qualquer), col11=Bloqueado
# df.iloc[4:, [0, 2, 6, 9, 11]]
# ─────────────────────────────────────────────────────────────────────────────
def mock_estoque_carga():
    wb = Workbook()
    wb.remove(wb.active)
    clientes = ['ADITUS', 'BIOGEN', 'IPSEN']
    skus_base = [
        ('ADI-001', 'PRODUTO ALFA 10MG', 100, 10, 0),
        ('ADI-002', 'PRODUTO BETA 20MG', 250, 25, 5),
        ('ADI-003', 'PRODUTO GAMA 50MG', 80, 0, 0),
        ('ADI-004', 'PRODUTO DELTA 100MG', 320, 32, 10),
        ('ADI-005', 'PRODUTO SIGMA 200MG', 50, 5, 0),
        ('ADI-006', 'PRODUTO OMEGA 5MG', 175, 17, 3),
    ]
    for cli in clientes:
        ws = wb.create_sheet(cli)
        # Linhas 0-3: cabeçalho do sistema (4 linhas ignoradas)
        ws.append(['SISTEMA', 'EXPORTACAO', 'ESTOQUE', date.today().strftime('%d/%m/%Y')])  # linha 1
        ws.append(['Empresa:', cli, '', '', '', '', '', '', '', '', '', ''])                 # linha 2
        ws.append(['Data Extração:', date.today().strftime('%d/%m/%Y')])                     # linha 3
        ws.append(['Código', 'UN', 'Descrição', 'Lote', 'Val', 'Local', 'Saldo',
                   'Fat', 'Valor', 'Reservado', 'Transit', 'Bloqueado'])                    # linha 4 (header)
        # Linha 4 = índice 4 no df (python 0-based) → dados a partir daqui
        for cod_base, desc, saldo, reservado, bloqueado in skus_base:
            cod = f'{cod_base[4:]}-{cli[:3]}'
            ws.append([cod, 'UN', desc, 'L001', '2027-12', 'EST-A',
                       saldo, 1000.0 + saldo * 10, saldo * 12.5,
                       reservado, 0, bloqueado])

    path = os.path.join(FIXTURES, 'mock_estoque_carga.xlsx')
    wb.save(path)
    print(f"✅ mock_estoque_carga.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 5. ESTOQUE — ATUALIZAR DB
# Abas por cliente.
# Linhas de movimentação: col0=data (DD/MM/YYYY), col2='[CÓDIGO] Descrição'
# Linhas de saldo final: col0='Saldo Final', col2='[CÓDIGO] Descrição', col11=saldo
# ─────────────────────────────────────────────────────────────────────────────
def mock_estoque_atualizar():
    wb = Workbook()
    wb.remove(wb.active)
    clientes = ['ADITUS', 'BIOGEN', 'IPSEN']
    today = date.today()

    for cli in clientes:
        ws = wb.create_sheet(cli)
        # Header mínimo (a função usa header=None e itera todas as linhas)
        sufixo = cli[:3]
        skus = [
            (f'001-{sufixo}', 'PRODUTO ALFA 10MG', 95, 92),
            (f'002-{sufixo}', 'PRODUTO BETA 20MG', 240, 238),
            (f'003-{sufixo}', 'PRODUTO GAMA 50MG', 80, 78),
        ]
        for cod, desc, saldo_inicial, saldo_final in skus:
            # Linha de saldo inicial
            row_ini = ['Saldo Inicial'] + [''] * 10 + [saldo_inicial] + [''] * 3 + [saldo_inicial * 12.5]
            row_ini[2] = f'[{cod}] {desc}'
            ws.append(row_ini)

            # 3 linhas de movimentação real
            for j in range(3):
                d = today - timedelta(days=j * 5)
                row = [d.strftime('%d/%m/%Y')] + [''] * 10 + [saldo_inicial - j * 2] + [''] * 3 + [(saldo_inicial - j * 2) * 12.5]
                row[2] = f'[{cod}] {desc}'
                ws.append(row)

            # Linha de saldo final (marcador que a função busca)
            row_fin = ['Saldo Final'] + [''] * 10 + [saldo_final] + [''] * 3 + [saldo_final * 12.5]
            row_fin[2] = f'[{cod}] {desc}'
            ws.append(row_fin)

    path = os.path.join(FIXTURES, 'mock_estoque_atualizar.xlsx')
    wb.save(path)
    print(f"✅ mock_estoque_atualizar.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 6. ESTOQUE — GERAR RELATÓRIO (arquivo de pico)
# Abas por cliente. Linhas 0-2: header. Linha 3+: dados diários.
# Última linha ignorada (dia da extração — incompleta).
# Col0=Data, Col10=Qtd, Col12=Valor, Col13=Área, Col14=Volume
# ─────────────────────────────────────────────────────────────────────────────
def mock_estoque_gerar_pico():
    wb = Workbook()
    wb.remove(wb.active)
    clientes = ['ADITUS', 'BIOGEN', 'IPSEN']
    today = date.today()
    mes_inicio = today.replace(day=1)

    for cli in clientes:
        ws = wb.create_sheet(cli)
        # 3 linhas de cabeçalho (linhas 0-2)
        ws.append(['SISTEMA DE ESTOQUE', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])  # linha 0
        ws.append(['Cliente:', cli, '', '', '', '', '', '', '', '', '', '', '', '', ''])           # linha 1
        ws.append(['Período:', 'Janeiro/2026', '', '', '', '', '', '', '', '', '', '', '', '', '']) # linha 2
        # Cabeçalho das colunas (linha 3, mas a função não usa header — accessa por índice)
        ws.append(['Data', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9',
                   'Qtd', 'C11', 'Valor', 'Area m2', 'Volume m3'])

        # Dados diários — ao menos 15 dias
        for day_offset in range(15):
            d = mes_inicio + timedelta(days=day_offset)
            qtd = 100 + day_offset * 5
            area = 50.0 + day_offset * 0.5
            volume = 20.0 + day_offset * 0.3
            valor = qtd * 12.5
            row = [d.strftime('%d/%m/%Y'), '', '', '', '', '', '', '', '', '',
                   qtd, '', valor, area, volume]
            ws.append(row)

        # Última linha (dia da extração — será ignorada pela função)
        ws.append([today.strftime('%d/%m/%Y'), '', '', '', '', '', '', '', '', '',
                   50, '', 625.0, 25.0, 10.0])

    path = os.path.join(FIXTURES, 'mock_estoque_pico.xlsx')
    wb.save(path)
    print(f"✅ mock_estoque_pico.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 8. RECEBIMENTOS
# Arquivo por índice (não por nome de coluna):
# col0=Seq, col1=(qualquer), col2=NF, col3..col5=(qualquer),
# col6=Depositante, col7=(qualquer), col8=Data, col9=Hora,
# col10..col13=(qualquer), col14=Fornecedor,
# col15..col18=(qualquer), col19=Valor,
# col20..col22=(qualquer), col23=Peso,
# col24=(qualquer), col25=Comentário, col26=Usuário
# Total de pelo menos 27 colunas
# ─────────────────────────────────────────────────────────────────────────────
def mock_recebimentos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Recebimentos"
    # Cabeçalho com 27 colunas (índices 0-26)
    header = [
        'Sequência', 'Tipo Doc.', 'Nota Fiscal', 'Série', 'Filial', 'Situação',
        'Depositante', 'Produto', 'Data Entrada', 'Hora Entrada',
        'C10', 'C11', 'C12', 'C13', 'Fornecedor',
        'C15', 'C16', 'C17', 'C18', 'Valor Nota',
        'C20', 'C21', 'C22', 'Peso',
        'C24', 'Comentário Adicional', 'Usuário Cadastro'
    ]
    ws.append(header)
    today = date.today()
    clientes = ['ADITUS', 'GSK', 'IPSEN', 'BIOGEN', 'CELLTRION']
    fornecedores = ['FORNEC A LTDA', 'FORNEC B SA', 'FORNEC C LTDA']

    nf_counter = 100
    for i in range(12):
        cli = clientes[i % len(clientes)]
        forn = fornecedores[i % len(fornecedores)]
        d = today - timedelta(days=30 - i * 2)

        # Linha normal (Entrada c/ NF)
        row = [''] * 27
        row[0] = i + 1
        row[2] = str(nf_counter)
        row[6] = cli
        row[8] = d.strftime('%d/%m/%Y')
        row[9] = '08:30'
        row[14] = forn
        row[19] = 1000.0 + i * 150
        row[23] = 50.0 + i * 5
        row[25] = ''
        row[26] = 'USUARIO.TESTE'
        ws.append(row)
        nf_counter += 1

    # Uma devolução
    row_dev = [''] * 27
    row_dev[0] = 99
    row_dev[2] = 'DEV-001'
    row_dev[6] = 'ADITUS'
    row_dev[8] = today.strftime('%d/%m/%Y')
    row_dev[9] = '14:00'
    row_dev[14] = 'FORNEC A LTDA'
    row_dev[19] = 500.0
    row_dev[23] = 20.0
    row_dev[25] = ''
    row_dev[26] = 'USUARIO.TESTE'
    ws.append(row_dev)

    # Um ajuste (deve ser ignorado)
    row_ajuste = [''] * 27
    row_ajuste[0] = 100
    row_ajuste[2] = 'AJUSTE'
    row_ajuste[6] = 'ADITUS'
    row_ajuste[8] = today.strftime('%d/%m/%Y')
    row_ajuste[9] = '10:00'
    row_ajuste[14] = ''
    row_ajuste[19] = 0
    row_ajuste[23] = 0
    row_ajuste[25] = ''
    row_ajuste[26] = 'SISTEMA'
    ws.append(row_ajuste)

    path = os.path.join(FIXTURES, 'mock_recebimentos.xlsx')
    wb.save(path)
    print(f"✅ mock_recebimentos.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 9. FATURAMENTO DISTRIBUIÇÃO
# Aba começando com 'EMBARQUES' (case insensitive).
# Colunas obrigatórias: Cliente Faturado, Valor Frete, Frete Peso, ADValorem,
# Gris, ITR, Despacho, SEC/CAT, Taxa ICMS, Valor ICMS, Outros, Valor N.F., Volume, Peso Taxado
# Filtra: remove linhas com Cliente Faturado vazio ou contendo 'BAIA 4'
# ─────────────────────────────────────────────────────────────────────────────
def mock_fat_dist():
    wb = Workbook()
    ws = wb.active
    ws.title = "EMBARQUES"
    # 'Centro de Custo' é acessado incondicionalmente em _fat_dist_aba_fechamento
    # (linha 5282 do central_relatorios.py) — coluna OBRIGATÓRIA mesmo sem clientes FAZBEM.
    # Nomes de clientes devem bater com _CC_CLIENTES: ADITUS HEALTH, BIOGEN BRASIL,
    # IPSEN FARMACEUTICA, YELUM SEGURADORA.
    header = [
        'Cliente Faturado', 'Centro de Custo', 'Valor Frete', 'Frete Peso', 'ADValorem',
        'Gris', 'ITR', 'Despacho', 'SEC/CAT', 'Taxa ICMS',
        'Valor ICMS', 'Outros', 'Valor N.F.', 'Volume', 'Peso Taxado'
    ]
    ws.append(header)
    dados = [
        ('ADITUS HEALTH',        'ADITUS_CC01',   2500.0, 500.0, 50.0, 30.0, 20.0, 10.0, 5.0, 0.12, 300.0, 15.0, 50000.0, 10.0, 80.0),
        ('BIOGEN BRASIL',        'MARKETING AF',  1800.0, 360.0, 36.0, 20.0, 15.0, 8.0,  4.0, 0.12, 216.0, 12.0, 36000.0, 8.0,  55.0),
        ('IPSEN FARMACEUTICA',   'IPSEN_CC01',    1200.0, 240.0, 24.0, 15.0, 10.0, 5.0,  3.0, 0.12, 144.0, 8.0,  24000.0, 5.0,  40.0),
        ('YELUM SEGURADORA',     'YELUM_CC01',     900.0, 180.0, 18.0, 12.0, 8.0,  4.0,  2.0, 0.0,    0.0, 6.0,  18000.0, 4.0,  30.0),
        ('ADITUS HEALTH',        'ADITUS_CC02',    600.0, 120.0, 12.0, 8.0,  5.0,  3.0,  1.5, 0.12,  72.0, 4.0,  12000.0, 2.5,  20.0),
        ('CELLTRION HEALTHCARE', '',               400.0,  80.0,  8.0, 5.0,  3.0,  2.0,  1.0, 0.12,  48.0, 3.0,   8000.0, 2.0,  15.0),
    ]
    for row in dados:
        ws.append(list(row))
    # Linha com cliente vazio (deve ser filtrada)
    ws.append(['', '', 100.0, 20.0, 2.0, 1.0, 0.5, 0.3, 0.2, 0.12, 12.0, 0.5, 2000.0, 0.5, 5.0])
    # Linha BAIA 4 (deve ser filtrada)
    ws.append(['BAIA 4 LOGISTICA', '', 200.0, 40.0, 4.0, 2.0, 1.0, 0.6, 0.4, 0.12, 24.0, 1.0, 4000.0, 1.0, 10.0])

    path = os.path.join(FIXTURES, 'mock_fat_dist.xlsx')
    wb.save(path)
    print(f"✅ mock_fat_dist.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 10a. FATURAMENTO ARMAZENAGEM — arquivo de volumes (pico por família)
# Abas por família/grupo/depositante único.
# Linha 3 (índice 3), col 0: filtro string: 'Familia do Produto = NOME; Per...'
# Linhas seguintes: col0=Data (DD/MM/YYYY), col12=Valor, col14=Volume m3
# ─────────────────────────────────────────────────────────────────────────────
def mock_fat_arm_volumes():
    wb = Workbook()
    wb.remove(wb.active)
    today = date.today()
    mes_inicio = today.replace(day=1)

    # Família de ADITUS
    ws1 = wb.create_sheet('ADITUS_FAM_A')
    ws1.append(['SISTEMA PICO', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])        # linha 0
    ws1.append(['Relatório de Volumes', '', '', '', '', '', '', '', '', '', '', '', '', '', '']) # linha 1
    ws1.append(['Filtros:', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])             # linha 2
    # Linha 3 - CRUCIAL: string de filtro com nome da família
    ws1.append([f'Familia do Produto = FAMILIA A; Per Inicial = {mes_inicio.strftime("%d/%m/%Y")}; Per Final = {today.strftime("%d/%m/%Y")}; Depositante = ADITUS',
               '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
    # Cabeçalho de colunas (linha 4 - não usada diretamente)
    ws1.append(['Data', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9',
                'C10', 'C11', 'Valor', 'C13', 'Volume m3'])
    # Dados diários (linhas 5+)
    for day_offset in range(10):
        d = mes_inicio + timedelta(days=day_offset)
        volume = 25.0 + day_offset * 0.8
        valor = volume * 150.0
        row = [d.strftime('%d/%m/%Y'), '', '', '', '', '', '', '', '', '', '', '',
               valor, '', volume]
        ws1.append(row)
    # Última linha (dia de extração — será ignorada)
    ws1.append([today.strftime('%d/%m/%Y'), '', '', '', '', '', '', '', '', '', '', '',
                3000.0, '', 20.0])

    # Família de BIOGEN
    ws2 = wb.create_sheet('BIOGEN_FAM_B')
    ws2.append(['SISTEMA PICO', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
    ws2.append(['Relatório de Volumes', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
    ws2.append(['Filtros:', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
    ws2.append([f'Familia do Produto = FAMILIA B; Per Inicial = {mes_inicio.strftime("%d/%m/%Y")}; Depositante = BIOGEN',
               '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
    ws2.append(['Data', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9',
                'C10', 'C11', 'Valor', 'C13', 'Volume m3'])
    for day_offset in range(10):
        d = mes_inicio + timedelta(days=day_offset)
        volume = 15.0 + day_offset * 0.5
        valor = volume * 130.0
        row = [d.strftime('%d/%m/%Y'), '', '', '', '', '', '', '', '', '', '', '',
               valor, '', volume]
        ws2.append(row)
    ws2.append([today.strftime('%d/%m/%Y'), '', '', '', '', '', '', '', '', '', '', '',
                1800.0, '', 12.0])

    path = os.path.join(FIXTURES, 'mock_fat_arm_volumes.xlsx')
    wb.save(path)
    print(f"✅ mock_fat_arm_volumes.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 10b. FATURAMENTO ARMAZENAGEM — arquivo de movimentação
# Abas por cliente. Mesmo formato que Estoque Atualizar:
# col0=data (DD/MM/YYYY) ou 'Saldo Inicial'/'Saldo Final'
# col2=produto '[COD] desc'
# col11=saldo_quantidade, col14=saldo_valor
# ─────────────────────────────────────────────────────────────────────────────
def mock_fat_arm_mov():
    wb = Workbook()
    wb.remove(wb.active)
    today = date.today()
    clientes = ['ADITUS', 'BIOGEN']

    for cli in clientes:
        ws = wb.create_sheet(cli)
        sufixo = cli[:3]
        skus = [
            (f'001-{sufixo}', 'PRODUTO ALFA'),
            (f'002-{sufixo}', 'PRODUTO BETA'),
            (f'003-{sufixo}', 'PRODUTO GAMA'),
        ]
        for cod, desc in skus:
            saldo_ini = 100
            saldo_fin = 90
            # Saldo inicial
            row_ini = ['Saldo Inicial'] + [''] * 14
            row_ini[2] = f'[{cod}] {desc}'
            row_ini[11] = saldo_ini
            row_ini[14] = saldo_ini * 12.5
            ws.append(row_ini)

            # Linhas de movimentação
            for j in range(5):
                d = today - timedelta(days=j * 3)
                row_mv = [d.strftime('%d/%m/%Y')] + [''] * 14
                row_mv[2] = f'[{cod}] {desc}'
                row_mv[11] = saldo_ini - j * 2
                row_mv[14] = (saldo_ini - j * 2) * 12.5
                ws.append(row_mv)

            # Saldo final
            row_fin = ['Saldo Final'] + [''] * 14
            row_fin[2] = f'[{cod}] {desc}'
            row_fin[11] = saldo_fin
            row_fin[14] = saldo_fin * 12.5
            ws.append(row_fin)

    path = os.path.join(FIXTURES, 'mock_fat_arm_mov.xlsx')
    wb.save(path)
    print(f"✅ mock_fat_arm_mov.xlsx — {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# DB DE ESTOQUE para testes (em vez de usar o DB de produção)
# ─────────────────────────────────────────────────────────────────────────────
def mock_db_estoque():
    import json
    db = {
        'ADITUS': {
            '001-ADI': {'desc': 'PRODUTO ALFA 10MG', 'saldo': 100, 'reservado': 10, 'bloqueado': 0, 'atualizado': date.today().strftime('%d/%m/%Y')},
            '002-ADI': {'desc': 'PRODUTO BETA 20MG', 'saldo': 250, 'reservado': 25, 'bloqueado': 5, 'atualizado': date.today().strftime('%d/%m/%Y')},
            '003-ADI': {'desc': 'PRODUTO GAMA 50MG', 'saldo': 80, 'reservado': 0, 'bloqueado': 0, 'atualizado': date.today().strftime('%d/%m/%Y')},
        },
        'BIOGEN': {
            '001-BIO': {'desc': 'PRODUTO ALFA 10MG', 'saldo': 50, 'reservado': 5, 'bloqueado': 0, 'atualizado': date.today().strftime('%d/%m/%Y')},
            '002-BIO': {'desc': 'PRODUTO BETA 20MG', 'saldo': 120, 'reservado': 12, 'bloqueado': 2, 'atualizado': date.today().strftime('%d/%m/%Y')},
        },
        'IPSEN': {
            '001-IPS': {'desc': 'PRODUTO DELTA 100MG', 'saldo': 200, 'reservado': 20, 'bloqueado': 0, 'atualizado': date.today().strftime('%d/%m/%Y')},
        }
    }
    path = os.path.join(FIXTURES, 'test_estoque_db.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    print(f"✅ test_estoque_db.json — {path}")
    return path


if __name__ == '__main__':
    print("\n=== Gerando mocks para Central de Relatórios ===\n")
    mock_pedidos()
    mock_fretes()
    mock_armazenagem()
    mock_estoque_carga()
    mock_estoque_atualizar()
    mock_estoque_gerar_pico()
    mock_recebimentos()
    mock_fat_dist()
    mock_fat_arm_volumes()
    mock_fat_arm_mov()
    mock_db_estoque()
    print(f"\n✅ Todos os mocks salvos em: {FIXTURES}\n")
    print("AVISO: Cap. Operacional (PDF) — SKIP. Requer pdfplumber + PDF real do Kardex.")
